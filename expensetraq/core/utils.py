from django.utils import six
from django.contrib.auth.decorators import user_passes_test
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.contrib.auth import REDIRECT_FIELD_NAME
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors, fonts
from datetime import timedelta
from io import BytesIO
import maya


def login_required(function=None, redirect_field_name=REDIRECT_FIELD_NAME, login_url=None):
    """
    Decorator for views that checks that the user is logged in, redirecting
    to the log-in page if necessary.
    """
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and u.is_active,
        login_url=login_url,
        redirect_field_name=redirect_field_name
    )
    if function:
        return actual_decorator(function)
    return actual_decorator


def user_in_groups(group, login_url=None, raise_exception=True):
    """
    Decorator for views that checks whether a user has a particular permission
    enabled, redirecting to the log-in page if necessary.
    If the raise_exception parameter is given the PermissionDenied exception
    is raised.
    """
    def check_groups(user):
        if isinstance(group, six.string_types):
            groups = (group, )
        else:
            groups = group
        # First check if the user has the permission (even anon users)
        if user.user_groups.intersection(set(groups)):
            return True
        # In case the 403 handler should be called raise the exception
        if raise_exception:
            raise PermissionDenied
        # As the last resort, show the login form
        return False

    return user_passes_test(check_groups, login_url=login_url)


class DeleteMessageMixin(object):
    """
    Adds a success message on successful form submission.
    """
    success_message = ''

    def delete(self, request, *args, **kwargs):
        success_message = self.get_success_message(self.get_object())
        response = super(DeleteMessageMixin, self).delete(request)
        if success_message:
            messages.success(self.request, success_message)
        return response

    def get_success_message(self, obj):
        return self.success_message % obj.__dict__


def context_processor(request):
    context = {}
    if request.user.is_authenticated:
        context['unread_notifications'] = request.user.notifications.filter(
            is_read=False)
    context['current_path'] = request.get_full_path()

    return context


def generate_expense_report(salesman, date_range, expense_list,
                            r_format='xls'):
    expense_report_stream = BytesIO()
    total_expenses = 0
    if r_format == 'xls':
        expense_report = Workbook()
        main_sheet = expense_report.active

        # Set up some initial styles
        main_sheet.freeze_panes = 'B2'
        main_sheet.row_dimensions[2].height = 32
        main_sheet.row_dimensions[3].height = 22
        main_sheet.row_dimensions[4].height = 22
        main_sheet.column_dimensions['A'].width = 22

        # Enter the name of the Salesman
        main_sheet['A2'] = 'Name'
        main_sheet['A2'].font = Font(
            name='Helvetica Neue', size=12, bold=True, color='FFFFFF')
        main_sheet['A2'].fill = PatternFill(
            start_color='004C7F', end_color='004C7F', fill_type='solid')
        main_sheet['B2'] = str(salesman)

        # Loop through the days and add the columns
        date_range_it = maya.MayaInterval(
            start=maya.when(date_range.split(' - ')[0]),
            end=maya.when(date_range.split(' - ')[1]).add(days=1)
        )

        col_idx = 2
        day_col_labels = {}
        all_col_labels = []
        for event in date_range_it.split(duration=timedelta(days=1)):
            col_label = get_column_letter(col_idx)
            day_col_labels[event.start.iso8601().split('T')[0]] = col_label
            # Apply styles to the columns
            main_sheet.column_dimensions[col_label].width = 16
            main_sheet[col_label + '2'].font = Font(
                name='Helvetica Neue', size=12, bold=True, color='FFFFFF')
            main_sheet[col_label + '2'].fill = PatternFill(
                start_color='004C7F', end_color='004C7F', fill_type='solid')

            main_sheet[col_label + '3'].font = Font(
                name='Helvetica Neue', size=11, bold=True, color='FFFFFF')
            main_sheet[col_label + '3'].fill = PatternFill(
                start_color='2F7115', end_color='2F7115', fill_type='solid')

            main_sheet[col_label + '3'] = \
                event.start.iso8601().split('T')[0]
            all_col_labels.append(col_label)
            col_idx += 1

        last_col_label = get_column_letter(col_idx)
        all_col_labels.append(last_col_label)

        # Set the date range in the sheet
        main_sheet[all_col_labels[-2] + '2'] = 'Date Range'
        main_sheet.column_dimensions[last_col_label].width = 16

        main_sheet[last_col_label + '2'] = date_range
        main_sheet[last_col_label + '2'].alignment = Alignment(
            wrap_text=True)
        main_sheet[last_col_label + '2'].font = Font(
            name='Helvetica Neue', size=12, bold=True, color='FFFFFF')
        main_sheet[last_col_label + '2'].fill = PatternFill(
            start_color='004C7F', end_color='004C7F', fill_type='solid')
        main_sheet[last_col_label + '3'] = '(Total)'
        main_sheet[last_col_label + '3'].font = Font(
            name='Helvetica Neue', size=11, bold=True, color='FFFFFF')
        main_sheet[last_col_label + '3'].fill = PatternFill(
            start_color='2F7115', end_color='2F7115', fill_type='solid')

        # Loop through the paid by and add it as a group
        cur_row = 4
        for paid_by in expense_list.keys():
            # Set the header for cash expenses
            main_sheet['A%s' % cur_row] = paid_by
            main_sheet['A%s' % cur_row].font = Font(
                name='Helvetica Neue', size=10, bold=True, color='FFFFFF')
            main_sheet['A%s' % cur_row].fill = PatternFill(
                start_color='004C7F', end_color='004C7F', fill_type='solid')

            for col_label in all_col_labels:
                main_sheet[col_label + str(cur_row)].font = Font(
                    name='Helvetica Neue', size=10, bold=True, color='FFFFFF')
                main_sheet[col_label + str(cur_row)].fill = PatternFill(
                    start_color='004C7F', end_color='004C7F', fill_type='solid')
            cur_row += 1

            # Add all the cash expenses here
            for expense_type, amounts in expense_list[paid_by].items():
                if expense_type != 'total':
                    main_sheet['A%s' % cur_row] = expense_type
                    main_sheet['A%s' % cur_row].font = Font(
                        name='Helvetica Neue', size=10, bold=True)
                    main_sheet['A%s' % cur_row].alignment = Alignment(
                        wrap_text=True)
                    line_amount = 0
                    for t_date, amount in amounts.items():
                        col_label = day_col_labels[t_date]
                        main_sheet[col_label + str(cur_row)] = amount
                        main_sheet[col_label + str(cur_row)].font = Font(
                            name='Helvetica Neue', size=10, bold=False)
                        line_amount += amount
                    main_sheet[last_col_label + str(cur_row)] = line_amount
                    main_sheet[last_col_label + str(cur_row)].font = Font(
                        name='Helvetica Neue', size=10, bold=True)
                    cur_row += 1

            # Set the trailer for cash expenses
            main_sheet.row_dimensions[cur_row].height = 22
            main_sheet['A%s' % cur_row] = 'Total %s Expenses' % paid_by
            main_sheet['A%s' % cur_row].font = Font(
                name='Helvetica Neue', size=10, bold=True)
            line_amount = 0
            for t_date, amount in expense_list[paid_by]['total'].items():
                col_label = day_col_labels[t_date]
                main_sheet[col_label + str(cur_row)] = amount
                main_sheet[col_label + str(cur_row)].font = Font(
                    name='Helvetica Neue', size=10, bold=True)
                line_amount += amount
            main_sheet[last_col_label + str(cur_row)] = line_amount
            main_sheet[last_col_label + str(cur_row)].font = Font(
                name='Helvetica Neue', size=10, bold=True)
            total_expenses += line_amount
            cur_row += 1

        # Set the trailer for the whole file
        main_sheet.row_dimensions[cur_row].height = 22
        main_sheet[all_col_labels[-2] + str(cur_row)] = 'Total:'
        main_sheet[all_col_labels[-2] + str(cur_row)].font = Font(
            name='Helvetica Neue', size=10, bold=True)
        main_sheet[last_col_label + str(cur_row)] = total_expenses
        main_sheet[last_col_label + str(cur_row)].font = Font(
            name='Helvetica Neue', size=10, bold=True)

        # Finally create a table for this
        wsprops = main_sheet.sheet_properties
        wsprops.pageSetUpPr = PageSetupProperties(
            fitToPage=True, autoPageBreaks=False)

        # Create the report and create the django response from it
        expense_report.save(expense_report_stream)

    elif r_format == 'pdf':
        expense_report = SimpleDocTemplate(
            expense_report_stream,
            rightMargin=72,
            leftMargin=72,
            topMargin=30,
            bottomMargin=72,
            pagesize=landscape(A4))

        # Loop through the days and add the columns
        date_range_it = maya.MayaInterval(
            start=maya.when(date_range.split(' - ')[0]),
            end=maya.when(date_range.split(' - ')[1]).add(days=1)
        )

        report_days = []
        for event in date_range_it.split(duration=timedelta(days=1)):
            report_days.append(event.start.iso8601().split('T')[0])
        total_columns = len(report_days) + 1

        report_data = [
            ['Name', str(salesman)] + [''] * (total_columns - 3) + ['Date Range', date_range],
            [''] + report_days + ['(Total)']
        ]
        report_style = [
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (total_columns, 0), '#004C7F'),
            ('FONTNAME', (0, 0), (total_columns, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (total_columns, 0), 10),
            ('TEXTCOLOR', (0, 0), (total_columns, 0), colors.white),
            ('BACKGROUND', (1, 1), (total_columns, 1), '#2F7115'),
            ('FONTNAME', (1, 1), (total_columns, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 1), (total_columns, 1), 9),
            ('TEXTCOLOR', (1, 1), (total_columns, 1), colors.white)]

        cur_row = 2
        for paid_by in expense_list.keys():
            # Add the header for this paid_by
            report_data.append([paid_by] + [''] * total_columns)
            report_style.extend([
                ('BACKGROUND', (0, cur_row), (total_columns, cur_row), '#004C7F'),
                ('FONTNAME', (0, cur_row), (total_columns, cur_row), 'Helvetica-Bold'),
                ('FONTSIZE', (0, cur_row), (total_columns, cur_row), 8),
                ('TEXTCOLOR', (0, cur_row), (total_columns, cur_row), colors.white)
            ])
            cur_row += 1
            # Add all the expenses for this paid_by
            for expense_type, amounts in expense_list[paid_by].items():
                if expense_type != 'total':
                    line = [expense_type]
                    line_amount = 0
                    for day in report_days:
                        line.append(amounts.get(day, ''))
                        line_amount += amounts.get(day, 0)
                    line.append(line_amount)
                    report_data.append(line)
                    report_style.extend([
                        ('FONTNAME', (0, cur_row), (0, cur_row), 'Helvetica-Bold'),
                        ('FONTNAME', (total_columns, cur_row), (total_columns, cur_row), 'Helvetica-Bold'),
                    ])
                    cur_row += 1

            # Add the trailer for this paid_by
            pd_trailer = ['Total %s Expenses' % paid_by]
            line_totals = 0
            for day in report_days:
                pd_trailer.append(expense_list[paid_by]['total'].get(day, ''))
                line_totals += expense_list[paid_by]['total'].get(day, 0)
            pd_trailer.append(line_totals)
            report_data.append(pd_trailer)
            report_style.append(
                ('FONTNAME', (0, cur_row), (total_columns, cur_row), 'Helvetica-Bold')
            )
            total_expenses += line_totals
            cur_row += 1

        # Add the report trailer
        report_data.append(
            [''] * (total_columns - 1) + ['Total:', total_expenses])
        report_style.append(
            ('FONTNAME', (0, cur_row), (total_columns, cur_row), 'Helvetica-Bold')
        )
        # Create the table and Set the style
        report_table = Table(report_data)
        report_table.setStyle(TableStyle(report_style))
        expense_report.build([report_table])

    return expense_report_stream.getvalue()
