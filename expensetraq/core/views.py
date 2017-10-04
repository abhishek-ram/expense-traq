# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.views.generic import TemplateView, ListView, CreateView, \
    UpdateView, DeleteView, DetailView, FormView
from django.contrib.messages.views import SuccessMessageMixin
from django.utils.decorators import method_decorator
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.views import View
from django.contrib import messages
from django.db import transaction
from django.forms import inlineformset_factory
from django.db.models import Sum
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from expensetraq.core.utils import user_in_groups, DeleteMessageMixin
from expensetraq.core.models import Expense, ExpenseType, \
    Salesman, ExpenseLimit, ExpenseLine, RecurringExpense, Notification, \
    CompanyCard, User, SalesmanCompanyCard, SalesmanExpenseType, Region
from expensetraq.core.forms import SalesmanForm, ExpenseLineForm, \
    ExpenseApprovalForm, ExpenseForm, DailyExpenseForm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import timedelta
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
import json
import maya

LIMIT_DATE = maya.when(timezone.now().isoformat()).subtract(
    months=3).datetime()


class Index(TemplateView):
    template_name = 'core/index.html'

    def get_context_data(self, **kwargs):
        context = super(Index, self).get_context_data(**kwargs)
        if self.request.user.is_admin:
            salesmen = Salesman.objects.all()
            context.update({
                'team_count': len(salesmen),
                'pending_amt': sum([
                    e.total_amount for e in Expense.objects.filter(status='P')
                ]),
                'approved_amt': sum([
                    e.total_amount for e in Expense.objects.filter(status='A')
                ]),
                'denied_amt': sum([
                    e.total_amount for e in Expense.objects.filter(status='D')
                ]),
                'salesman_list': salesmen,
                'expense_list': Expense.objects.all()[:10]
            })

        elif self.request.user.is_salesman:
            if hasattr(self.request.user, 'salesman'):
                expenses = self.request.user.salesman.expenses.filter(
                    transaction_date__gte=LIMIT_DATE)
                context.update({
                    'pending_amt': sum([
                        e.total_amount for e in expenses.filter(status='P')]),
                    'approved_amt': sum([
                        e.total_amount for e in expenses.filter(status='A')]),
                    'denied_amt': sum([
                        e.total_amount for e in expenses.filter(status='D')]),
                    'expense_list': expenses[:10],
                    'daily_form': DailyExpenseForm(
                        salesman=self.request.user.salesman)
                })
        elif self.request.user.is_manager:
            team = self.request.user.team.all()
            expenses = Expense.objects.filter(
                transaction_date__gte=LIMIT_DATE, salesman__in=team)
            context.update({
                'team_count': len(team),
                'pending_amt': sum([
                    e.total_amount for e in expenses.filter(status='P')]),
                'approved_amt': sum([
                    e.total_amount for e in expenses.filter(status='A')]),
                'denied_amt': sum([
                    e.total_amount for e in expenses.filter(status='D')]),
                'expense_list': expenses[:10]
            })

        return context


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseTypeList(ListView):
    model = ExpenseType


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseTypeCreate(CreateView):
    model = ExpenseType
    fields = '__all__'
    success_url = reverse_lazy('expense-type-list')
    success_message = 'Expense Type "%(name)s" has been created successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseTypeUpdate(UpdateView):
    model = ExpenseType
    fields = '__all__'
    success_url = reverse_lazy('expense-type-list')
    success_message = 'Expense Type "%(name)s" has been edited successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseTypeDelete(DeleteMessageMixin, DeleteView):
    model = ExpenseType
    success_url = reverse_lazy('expense-type-list')
    success_message = 'Expense Type "%(name)s" has been deleted successfully'


@method_decorator(user_in_groups(['Expense-User',
                                  'Expense-Admin']), name='dispatch')
class ExpenseCreate(CreateView):
    model = Expense
    form_class = ExpenseForm
    success_message = 'Expense with total ${0.total_amount} has ' \
                      'been recorded successfully'
    ExpenseFormSet = inlineformset_factory(
        Expense, ExpenseLine, extra=4, min_num=1, can_delete=False,
        validate_min=True, form=ExpenseLineForm)

    def get_context_data(self, **kwargs):
        context = super(ExpenseCreate, self).get_context_data(**kwargs)
        if self.request.user.is_admin:
            context['salesman'] = Salesman.objects.get(
                id=self.request.GET.get('salesman'))
        else:
            context['salesman'] = self.request.user.salesman
        expense_types = {}
        for et in context['salesman'].expense_types\
                .exclude(expense_type__name__in=['Daily Expense']):
            if not expense_types.get(et.region.id):
                expense_types[et.region.id] = []
            expense_types[et.region.id].append([et.id, et.expense_type.name])
        context['expense_types'] = json.dumps(expense_types)

        if not context.get('inline_formset'):
            context['inline_formset'] = self.ExpenseFormSet(
                form_kwargs={'salesman': context['salesman'],
                             'user_is_admin': self.request.user.is_admin})
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                # Save the expense object
                self.object = form.save()

                # Initialise the expense line formset
                formset = self.ExpenseFormSet(
                    self.request.POST, instance=self.object,
                    form_kwargs={'salesman': self.object.salesman,
                                 'user_is_admin': self.request.user.is_admin})

                # Validate and save the formset
                assert formset.is_valid()
                formset.save()
                messages.success(
                    self.request, self.get_success_message(self.object))

                # Create notifications for the admin and manager
                not_message = '{} has added a new expense dated {} for ${}'.format(
                    self.request.user, self.object.transaction_date,
                    self.object.total_amount)
                if self.object.salesman.manager:
                    Notification.objects.create(
                        user=self.object.salesman.manager,
                        title='New Expense Created',
                        text=not_message)
                Notification.objects.create(
                    user=User.objects.filter(
                        groups__name__in=['Expense-Admin']).first(),
                    title='New Expense Created',
                    text=not_message)
                return HttpResponseRedirect(self.get_success_url())
        except AssertionError:
            form.instance.id = None
            messages.error(
                self.request, 'Failed to record expense, correct errors and resubmit')
            return self.render_to_response(
                self.get_context_data(form=form, inline_formset=formset))

    def get_success_url(self):
        if self.request.user.is_admin:
            return reverse_lazy('expense-approval')
        else:
            return '%s?action=list' % reverse_lazy('expense-list-export')

    def get_success_message(self, obj):
        return self.success_message.format(obj)


@method_decorator(user_in_groups(['Expense-Admin', 'Expense-User']),
                  name='dispatch')
class ExpenseUpdate(UpdateView):
    model = Expense
    fields = ['status', 'transaction_date', 'paid_by', 'notes']
    success_message = 'Expense "{0.pk}" has been edited successfully'

    ExpenseFormSet = inlineformset_factory(
        Expense, ExpenseLine, max_num=5, can_delete=False, form=ExpenseLineForm)

    def get_context_data(self, **kwargs):
        context = super(ExpenseUpdate, self).get_context_data(**kwargs)
        if not context.get('inline_formset'):
            context['inline_formset'] = self.ExpenseFormSet(
                form_kwargs={'salesman': context['form'].instance.salesman,
                             'user_is_admin': self.request.user.is_admin},
                instance=context['form'].instance)
        context['salesman'] = context['form'].instance.salesman

        expense_types = {}
        for et in context['salesman'].expense_types \
                .exclude(expense_type__name__in=['Daily Expense']):
            if not expense_types.get(et.region.id):
                expense_types[et.region.id] = []
            expense_types[et.region.id].append([et.id, et.expense_type.name])
        context['expense_types'] = json.dumps(expense_types)

        if not self.request.user.is_admin \
                and context['salesman'] != self.request.user.salesman:
            raise PermissionDenied
        return context

    def form_valid(self, form):
        try:
            self.object = form.save()
            formset = self.ExpenseFormSet(
                self.request.POST, instance=self.object,
                form_kwargs={'salesman': self.object.salesman,
                             'user_is_admin': self.request.user.is_admin})
            assert formset.is_valid()
            formset.save()

            # Create notifications for the salesman and manager
            if self.request.user.salesman != self.object.salesman:
                Notification.objects.create(
                    user=self.object.salesman.user,
                    title='Expense Updated',
                    text='Admin has edited your expense dated {} for ${}'.format(
                        self.object.transaction_date, self.object.total_amount))
            else:
                Notification.objects.create(
                    user=User.objects.filter(
                        groups__name__in=['Expense-Admin']).first(),
                    title='Expense Updated',
                    text='{}\'s has edited his expense dated {} for ${}'.format(
                        self.object.salesman.user, self.object.transaction_date,
                        self.object.total_amount))

            messages.success(self.request,
                             self.get_success_message(self.object))
            return HttpResponseRedirect(self.get_success_url())
        except AssertionError:
            messages.error(
                self.request,
                'Failed to record expense, correct errors and resubmit')
            return self.render_to_response(
                self.get_context_data(form=form, inline_formset=formset))

    def get_success_message(self, obj):
        return self.success_message.format(obj)

    def get_success_url(self):
        if self.request.user.is_admin:
            return reverse_lazy('expense-approval')
        else:
            print('%s?action=list' % reverse_lazy('expense-list-export'))
            return '%s?action=list' % reverse_lazy('expense-list-export')


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseApproval(ListView):
    model = Expense
    form_class = ExpenseApprovalForm
    template_name = 'core/expense_approval.html'

    def get_queryset(self):
        # salesman = self.request.GET.get('salesman')
        return Expense.objects.filter(status='P')

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            ap_display = 'Approved' if form.cleaned_data['approved'] else 'Denied'
            if form.cleaned_data['approved']:
                form.cleaned_data['expense_list'].update(status='A')
            else:
                form.cleaned_data['expense_list'].update(status='D')
            messages.success(
                request, 'Selected expenses have been %s' % ap_display)

            # Create notifications for the salesman and manager
            for expense in form.cleaned_data['expense_list']:
                Notification.objects.create(
                    user=expense.salesman.user,
                    title='Expense Updated',
                    text='Admin has {} your expense dated {} for ${}'.format(
                        ap_display, expense.transaction_date, expense.total_amount))
                Notification.objects.create(
                    user=expense.salesman.manager,
                    title='Expense Updated',
                    text='Admin has {} {}\'s expense dated {} for ${}'.format(
                        ap_display, expense.salesman.user,
                        expense.transaction_date, expense.total_amount))

        else:
            messages.error(
                request, 'Unable to update expenses, please contact Admin')
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        """
        Insert the form into the context dict.
        """
        return super(ExpenseApproval, self).get_context_data(**kwargs)

    def get_form_kwargs(self):
        """
        Returns the keyword arguments for instantiating the form.
        """
        kwargs = {
            # 'initial': self.get_initial(),
            # 'prefix': self.get_prefix(),
            # 'user': self.request.user,
        }

        if self.request.method in ('POST', 'PUT'):
            kwargs.update({
                'data': self.request.POST,
                'files': self.request.FILES,
            })
        else:
            kwargs.update({
                'data': self.request.GET,
            })
        return kwargs

    def get_form(self):
        return self.form_class(**self.get_form_kwargs())


class ExpenseDetail(DetailView):
    model = Expense


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseDelete(DeleteMessageMixin, DeleteView):
    model = Expense
    success_message = 'Expense <var>%(id)s</var> has been deleted ' \
                      'successfully'

    def get_success_url(self):
        return self.request.POST['success_url']


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class SalesmanList(ListView):
    model = Salesman


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class SalesmanCreate(SuccessMessageMixin, CreateView):
    model = Salesman
    form_class = SalesmanForm
    success_url = reverse_lazy('salesman-list')
    success_message = 'Salesman "%(user)s" has been added successfully'
    CCFormSet = inlineformset_factory(
        Salesman, SalesmanCompanyCard, extra=2, can_delete=False,
        fields=['company_card', 'gp_vendor_code'])
    ETFormSet = inlineformset_factory(
        Salesman, SalesmanExpenseType, extra=5, can_delete=False,
        fields=['expense_type', 'region', 'gl_code_suffix'])

    def get_context_data(self, **kwargs):
        context = super(SalesmanCreate, self).get_context_data(**kwargs)
        if not context.get('cc_formset'):
            context['cc_formset'] = self.CCFormSet()
        if not context.get('et_formset'):
            context['et_formset'] = self.ETFormSet()
        context['expense_prefixes'] = json.dumps(
            {et.id: et.gl_code_prefix for et in ExpenseType.objects.all()})
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                self.object = form.save()
                cc_formset = self.CCFormSet(
                    self.request.POST, instance=self.object)
                et_formset = self.ETFormSet(
                    self.request.POST, instance=self.object)

                # Save the credit cards for the salesman
                assert cc_formset.is_valid()
                cc_formset.save()

                # Save the expense types for the salesman
                assert et_formset.is_valid()
                et_formset.save()

                messages.success(self.request,
                                 self.get_success_message(form.cleaned_data))
                return HttpResponseRedirect(self.get_success_url())
        except AssertionError:
            if cc_formset.errors:
                messages.error(
                    self.request, 'Failed to save Salesman, Errors are '
                                  '{}'.format(cc_formset.errors[0].as_ul())
                )
            else:
                messages.error(
                    self.request, 'Failed to save Salesman, Errors are '
                                  '{}'.format(et_formset.errors[0].as_ul())
                )
            return self.render_to_response(
                self.get_context_data(
                    form=form, cc_formset=cc_formset, et_formset=et_formset))

    def get_success_message(self, cleaned_data):
        return self.success_message % cleaned_data


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class SalesmanUpdate(SuccessMessageMixin, UpdateView):
    model = Salesman
    form_class = SalesmanForm
    success_url = reverse_lazy('salesman-list')
    success_message = 'Salesman "%(user)s" has been edited successfully'

    CCFormSet = inlineformset_factory(
        Salesman, SalesmanCompanyCard, extra=1,
        fields=['company_card', 'gp_vendor_code'])

    ETFormSet = inlineformset_factory(
        Salesman, SalesmanExpenseType, extra=3,
        fields=['expense_type', 'region', 'gl_code_suffix'])

    def get_context_data(self, **kwargs):
        context = super(SalesmanUpdate, self).get_context_data(**kwargs)
        if not context.get('cc_formset'):
            context['cc_formset'] = self.CCFormSet(
                instance=context['form'].instance)
        if not context.get('et_formset'):
            context['et_formset'] = self.ETFormSet(
                instance=context['form'].instance)
        context['expense_prefixes'] = json.dumps(
            {et.id: et.gl_code_prefix for et in ExpenseType.objects.all()})
        return context

    def form_valid(self, form):
        self.object = form.save()
        # Save the company card and expense types for the salesman
        cc_formset = self.CCFormSet(self.request.POST, instance=self.object)
        et_formset = self.ETFormSet(self.request.POST, instance=self.object)
        if cc_formset.is_valid() and et_formset.is_valid():
            cc_formset.save()
            et_formset.save()
            messages.success(
                self.request, self.get_success_message(form.cleaned_data))
            return HttpResponseRedirect(self.get_success_url())
        else:
            if cc_formset.errors:
                messages.error(
                    self.request, 'Failed to save Salesman, Errors are '
                                  '{}'.format(cc_formset.errors[0].as_ul())
                )
            else:
                messages.error(
                    self.request, 'Failed to save Salesman, Errors are '
                                  '{}'.format(et_formset.errors[0].as_ul())
                )
            return self.render_to_response(
                self.get_context_data(
                    form=form, cc_formset=cc_formset, et_formset=et_formset))


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseLimitList(ListView):
    model = ExpenseLimit


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseLimitCreate(SuccessMessageMixin, CreateView):
    model = ExpenseLimit
    fields = '__all__'
    success_url = reverse_lazy('expense-limit-list')
    success_message = 'Expense limit for <var>%(salesman)s</var> has been ' \
                      'added successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseLimitUpdate(SuccessMessageMixin, UpdateView):
    model = ExpenseLimit
    fields = '__all__'
    success_url = reverse_lazy('expense-limit-list')
    success_message = 'Expense limit for <var>%(salesman)s</var> has been ' \
                      'edited successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class ExpenseLimitDelete(DeleteMessageMixin, DeleteView):
    model = ExpenseLimit
    success_url = reverse_lazy('expense-limit-list')
    success_message = 'Expense Type <var>%(id)s</var> has been deleted ' \
                      'successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class RecurringExpenseList(ListView):
    model = RecurringExpense


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class RecurringExpenseCreate(SuccessMessageMixin, CreateView):
    model = RecurringExpense
    fields = '__all__'
    success_url = reverse_lazy('recur-expense-list')
    success_message = 'Recurring Expense for <var>%(salesman)s</var> has ' \
                      'been added successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class RecurringExpenseUpdate(SuccessMessageMixin, UpdateView):
    model = RecurringExpense
    fields = '__all__'
    success_url = reverse_lazy('recur-expense-list')
    success_message = 'Recurring Expense for <var>%(salesman)s</var> has ' \
                      'been edited successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class RecurringExpenseDelete(DeleteMessageMixin, DeleteView):
    model = RecurringExpense
    success_url = reverse_lazy('recur-expense-list')
    success_message = 'Recurring Expense <var>%(id)s</var> has been deleted ' \
                      'successfully'


class NotificationList(ListView):
    model = Notification
    paginate_by = 10

    def get_queryset(self):
        return self.request.user.notifications.all()


class ExpenseDailyAverage(View):

    def get(self, request):
        series = []
        for date in maya.MayaInterval(duration=30*24*60*60, end=maya.now().add(
                days=1)).split(duration=24*60*60):
            daily_expense = request.user.salesman.expenses.\
                filter(transaction_date=date.start.datetime().date(),
                       status='A').\
                aggregate(Sum('lines__amount'))['lines__amount__sum'] or 0
            series.append([date.start.datetime().strftime('%Y-%m-%d'),
                           daily_expense])
        return JsonResponse(series, safe=False)


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class CompanyCardList(ListView):
    model = CompanyCard


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class CompanyCardCreate(SuccessMessageMixin, CreateView):
    model = CompanyCard
    fields = '__all__'
    success_url = reverse_lazy('company-card-list')
    success_message = 'Company Card <var>%(name)s</var> has ' \
                      'been added successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class CompanyCardUpdate(SuccessMessageMixin, UpdateView):
    model = CompanyCard
    fields = '__all__'
    success_url = reverse_lazy('company-card-list')
    success_message = 'Company Card <var>%(name)s</var> has ' \
                      'been edited successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class CompanyCardDelete(DeleteMessageMixin, DeleteView):
    model = CompanyCard
    success_url = reverse_lazy('company-card-list')
    success_message = 'Company Card <var>%(name)s</var> has been deleted ' \
                      'successfully'


class ExpenseListExport(ListView):
    model = Expense

    def get_queryset(self):
        qs = Expense.objects.filter(transaction_date__gte=LIMIT_DATE)

        # maintain a list of salesman to be used in the export
        salesman_list = self.request.GET.getlist('salesman[]')
        if salesman_list:
            qs = qs.filter(salesman_id__in=salesman_list)
        else:
            if self.request.user.is_salesman and not self.request.user.is_admin:
                qs = qs.filter(salesman_id=self.request.user.salesman)
            elif self.request.user.is_manager:
                qs.filter(
                    salesman_id__in=[s.id for s in self.request.user.team.all()])

        status_list = self.request.GET.getlist('status[]')
        if status_list:
            qs = qs.filter(status__in=status_list)

        paid_by_list = self.request.GET.getlist('paid_by[]')
        if paid_by_list:
            qs = qs.filter(paid_by__in=paid_by_list)

        date_range = self.request.GET.get('daterange')
        if date_range:
            qs = qs.filter(transaction_date__gte=date_range.split(' - ')[0]).\
                filter(transaction_date__lte=date_range.split(' - ')[1])

        return qs

    def get_context_data(self, **kwargs):
        context = super(ExpenseListExport, self).get_context_data(**kwargs)
        if self.request.user.is_manager:
            context['salesman_list'] = self.request.user.team.all()
        elif self.request.user.is_admin:
            context['salesman_list'] = Salesman.objects.all()
        context['status_list'] = Expense.STATUS_CHOICES
        context['paid_by_list'] = ['Employee Paid'] + \
                                  [cc.name for cc in CompanyCard.objects.all()]
        if self.request.GET.get('action') == 'list':
            context['total_amount'] = sum(
                [e.total_amount for e in self.object_list])
            context['all_status'] = {
                e.get_status_display() for e in self.object_list}
            context['all_paid_by'] = {
                e.paid_by for e in self.object_list}
            context['all_salesman'] = {
                str(e.salesman) for e in self.object_list}
            expense_types = []
            for expense in self.object_list:
                expense_types.extend(
                    [el.expense_type for el in expense.lines.all()])
            context['all_expense_type'] = set(expense_types)
        return context

    def get(self, request, *args, **kwargs):
        if self.request.GET.get('action') == 'export':
            salesman_expenses = {}
            date_range = self.request.GET.get('daterange')
            # Initialize the array of salesman
            for salesman in self.request.GET.getlist('salesman[]'):
                salesman_expenses[int(salesman)] = {}

            # Loop through the expense types and sort it
            for et in self.get_queryset().all():
                if not salesman_expenses[et.salesman.id].get(et.paid_by):
                    salesman_expenses[et.salesman.id][et.paid_by] = {
                        'total': {}
                    }

                trans_date = et.transaction_date.strftime('%Y-%m-%d')
                for line in et.lines.all():
                    e_type = str(line.expense_type)

                    if not salesman_expenses[et.salesman.id][
                            et.paid_by].get(e_type):
                        salesman_expenses[et.salesman.id][
                            et.paid_by][e_type] = {}
                    if not salesman_expenses[et.salesman.id][
                            et.paid_by][e_type].get(trans_date):
                        salesman_expenses[et.salesman.id][
                            et.paid_by][e_type][trans_date] = 0
                    if not salesman_expenses[et.salesman.id][
                            et.paid_by]['total'].get(trans_date):
                        salesman_expenses[et.salesman.id][
                            et.paid_by]['total'][trans_date] = 0
                    salesman_expenses[et.salesman.id][
                        et.paid_by][e_type][trans_date] += line.amount
                    salesman_expenses[et.salesman.id][
                        et.paid_by]['total'][trans_date] += line.amount

            expense_report_files = {}
            for sid, expense_list in salesman_expenses.items():                
                salesman = Salesman.objects.get(pk=sid)
                total_expenses = 0
                # Load the template report
                expense_report = Workbook()
                main_sheet = expense_report.active
    
                # Set up some initial styles
                main_sheet.freeze_panes = 'B2'
                main_sheet.row_dimensions[2].height = 32
                main_sheet.row_dimensions[3].height = 22
                main_sheet.row_dimensions[4].height = 22
                main_sheet.column_dimensions['A'].width = 22
    
                # Enter the name of the Salesman
                main_sheet['A2'] = 'Name'
                main_sheet['A2'].font = Font(
                    name='Helvetica Neue', size=12, bold=True, color='FFFFFF')
                main_sheet['A2'].fill = PatternFill(
                    start_color='004C7F', end_color='004C7F', fill_type='solid')
                main_sheet['B2'] = str(salesman)
    
                # Loop through the days and add the columns
                date_range_it = maya.MayaInterval(
                        start=maya.when(date_range.split(' - ')[0]),
                        end=maya.when(date_range.split(' - ')[1]).add(days=1)
                )
    
                col_idx = 2
                day_col_labels = {}
                all_col_labels = []
                for event in date_range_it.split(duration=timedelta(days=1)):
                    col_label = get_column_letter(col_idx)
                    day_col_labels[event.start.iso8601().split('T')[0]] = col_label
                    # Apply styles to the columns
                    main_sheet.column_dimensions[col_label].width = 16
                    main_sheet[col_label + '2'].font = Font(
                        name='Helvetica Neue', size=12, bold=True, color='FFFFFF')
                    main_sheet[col_label + '2'].fill = PatternFill(
                        start_color='004C7F', end_color='004C7F', fill_type='solid')
    
                    main_sheet[col_label + '3'].font = Font(
                        name='Helvetica Neue', size=11, bold=True, color='FFFFFF')
                    main_sheet[col_label + '3'].fill = PatternFill(
                        start_color='2F7115', end_color='2F7115', fill_type='solid')
    
                    main_sheet[col_label + '3'] =\
                        event.start.iso8601().split('T')[0]
                    all_col_labels.append(col_label)
                    col_idx += 1

                last_col_label = get_column_letter(col_idx)
                all_col_labels.append(last_col_label)

                # Set the date range in the sheet
                main_sheet[all_col_labels[-2] + '2'] = 'Date Range'
                main_sheet.column_dimensions[last_col_label].width = 16
    
                main_sheet[last_col_label + '2'] = date_range
                main_sheet[last_col_label + '2'].alignment = Alignment(
                    wrap_text=True)
                main_sheet[last_col_label + '2'].font = Font(
                    name='Helvetica Neue', size=12, bold=True, color='FFFFFF')
                main_sheet[last_col_label + '2'].fill = PatternFill(
                    start_color='004C7F', end_color='004C7F', fill_type='solid')
                main_sheet[last_col_label + '3'] = '(Total)'
                main_sheet[last_col_label + '3'].font = Font(
                    name='Helvetica Neue', size=11, bold=True, color='FFFFFF')
                main_sheet[last_col_label + '3'].fill = PatternFill(
                    start_color='2F7115', end_color='2F7115', fill_type='solid')

                # Loop through the paid by and add it as a group
                cur_row = 4
                for paid_by in expense_list.keys():
                    # Set the header for cash expenses
                    main_sheet['A%s' % cur_row] = paid_by
                    main_sheet['A%s' % cur_row].font = Font(
                        name='Helvetica Neue', size=10, bold=True, color='FFFFFF')
                    main_sheet['A%s' % cur_row].fill = PatternFill(
                        start_color='004C7F', end_color='004C7F', fill_type='solid')

                    for col_label in all_col_labels:
                        main_sheet[col_label + str(cur_row)].font = Font(
                            name='Helvetica Neue', size=10, bold=True, color='FFFFFF')
                        main_sheet[col_label + str(cur_row)].fill = PatternFill(
                            start_color='004C7F', end_color='004C7F', fill_type='solid')
                    cur_row += 1

                    # Add all the cash expenses here
                    for expense_type, amounts in expense_list[paid_by].items():
                        if expense_type != 'total':
                            main_sheet['A%s' % cur_row] = expense_type
                            main_sheet['A%s' % cur_row].font = Font(
                                name='Helvetica Neue', size=10, bold=True)
                            main_sheet['A%s' % cur_row].alignment = Alignment(
                                wrap_text=True)
                            line_amount = 0
                            for t_date, amount in amounts.items():
                                col_label = day_col_labels[t_date]
                                main_sheet[col_label + str(cur_row)] = amount
                                main_sheet[col_label + str(cur_row)].font = Font(
                                    name='Helvetica Neue', size=10, bold=False)
                                line_amount += amount
                            main_sheet[last_col_label + str(cur_row)] = line_amount
                            main_sheet[last_col_label + str(cur_row)].font = Font(
                                name='Helvetica Neue', size=10, bold=True)
                            cur_row += 1

                    # Set the trailer for cash expenses
                    main_sheet.row_dimensions[cur_row].height = 22
                    main_sheet['A%s' % cur_row] = 'Total Cash Expenses'
                    main_sheet['A%s' % cur_row].font = Font(
                        name='Helvetica Neue', size=10, bold=True)
                    line_amount = 0
                    for t_date, amount in expense_list[
                            paid_by]['total'].items():
                        col_label = day_col_labels[t_date]
                        main_sheet[col_label + str(cur_row)] = amount
                        main_sheet[col_label + str(cur_row)].font = Font(
                            name='Helvetica Neue', size=10, bold=True)
                        line_amount += amount
                    main_sheet[last_col_label + str(cur_row)] = line_amount
                    main_sheet[last_col_label + str(cur_row)].font = Font(
                        name='Helvetica Neue', size=10, bold=True)
                    total_expenses += line_amount
                    cur_row += 1

                # Set the trailer for the whole file
                main_sheet.row_dimensions[cur_row].height = 22
                main_sheet[all_col_labels[-2] + str(cur_row)] = 'Total:'
                main_sheet[all_col_labels[-2] + str(cur_row)].font = Font(
                    name='Helvetica Neue', size=10, bold=True)
                main_sheet[last_col_label + str(cur_row)] = total_expenses
                main_sheet[last_col_label + str(cur_row)].font = Font(
                    name='Helvetica Neue', size=10, bold=True)

                # Finally create a table for this
                from openpyxl.worksheet.properties import WorksheetProperties, \
                    PageSetupProperties
                wsprops = main_sheet.sheet_properties
                wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True,
                                                          autoPageBreaks=False)

                # Create the report and create the django response from it
                out_stream = BytesIO()
                expense_report.save(out_stream)

                filename = 'ExpenseReport_%s_%s.xlsx' % (salesman, date_range)
                expense_report_files[filename] = out_stream.getvalue()

            file_names = list(expense_report_files.keys())
            if len(file_names) == 1:
                filename = file_names[0]
                response = HttpResponse(
                    expense_report_files[filename],
                    content_type='application/vnd.openxmlformats-officedocument.'
                                 'spreadsheetml.sheet')
                response['Content-Disposition'] = \
                    'attachment; filename="%s"' % filename
            else:
                zip_stream = BytesIO()
                with ZipFile(zip_stream, mode='w',
                             compression=ZIP_DEFLATED) as zf:
                    for filename, content in expense_report_files.items():
                        zf.writestr(filename, content)
                response = HttpResponse(
                    zip_stream.getvalue(),
                    content_type='application/x-zip-compressed')
                response['Content-Disposition'] = \
                    'attachment; filename="ExpenseReports_%s.zip"' % date_range

            return response
        else:
            return super(ExpenseListExport, self).get(request, *args, **kwargs)


@method_decorator(user_in_groups(['Expense-User']), name='dispatch')
class DailyExpenseSubmit(FormView):
    form_class = DailyExpenseForm
    success_url = reverse_lazy('index')

    def get_form_kwargs(self):
        kwargs = super(DailyExpenseSubmit, self).get_form_kwargs()
        kwargs['salesman'] = self.request.user.salesman
        return kwargs

    def form_invalid(self, form):
        messages.error(
            self.request,
            'Daily Expense was not logged, Please select an expense type')
        return HttpResponseRedirect(reverse_lazy('index'))

    @transaction.atomic
    def form_valid(self, form):
        # Get the list of daily expenses logged by the user
        exist_expense = form.salesman.expenses.filter(
            transaction_date=form.cleaned_data['transaction_date'],
            lines__expense_type=form.cleaned_data['expense_type']
        ).first()

        # Calculate expense amount based on the worked hours
        if form.cleaned_data['worked'] == 'Full':
            expense_amount = form.salesman.daily_expense
        elif form.cleaned_data['worked'] == 'Half':
            expense_amount = form.salesman.daily_expense / 2
        elif form.cleaned_data['worked'] == 'Quart':
            expense_amount = form.salesman.daily_expense / 4

        # Create a new daily if no expense has been logged for that day or
        # half day has been logged
        if not exist_expense:
            expense = Expense.objects.create(
                salesman=form.salesman,
                transaction_date=form.cleaned_data['transaction_date'],
                paid_by='Employee Paid',
                notes=''
            )
            ExpenseLine.objects.create(
                expense=expense,
                expense_type=form.cleaned_data['expense_type'],
                amount=expense_amount
            )
            messages.success(
                self.request,
                '{} Daily Expense has been successfully logged '
                'for date {}'.format(form.cleaned_data['worked'],
                                     form.cleaned_data['transaction_date']))
        elif (exist_expense.total_amount + expense_amount) \
                <= form.salesman.daily_expense:
            line = exist_expense.lines.all().first()
            line.amount = exist_expense.total_amount + expense_amount
            line.save()
            messages.success(
                self.request,
                'Full Daily Expense has been successfully logged '
                'for date {}'.format(form.cleaned_data['transaction_date']))
        else:
            messages.error(
                self.request,
                'Cannot log Daily Expense as expense has already been logged '
                'for date {}'.format(form.cleaned_data['transaction_date']))

        return super(DailyExpenseSubmit, self).form_valid(form)


class SalesmanExpenseTypeList(View):

    def get(self, request, salesman_id):
        expense_types = [['', '---------']]
        all_expenses = SalesmanExpenseType.objects.\
            filter(salesman=salesman_id).\
            exclude(expense_type__name__in=['Daily Expense'])
        for et in all_expenses:
            expense_types.append([et.id, str(et)])
        return JsonResponse(expense_types, safe=False)


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class RegionList(ListView):
    model = Region


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class RegionCreate(CreateView):
    model = Region
    fields = '__all__'
    success_url = reverse_lazy('region-list')
    success_message = 'Region "%(name)s" has been created successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class RegionUpdate(UpdateView):
    model = Region
    fields = '__all__'
    success_url = reverse_lazy('region-list')
    success_message = 'Region "%(name)s" has been edited successfully'


@method_decorator(user_in_groups(['Expense-Admin']), name='dispatch')
class RegionDelete(DeleteMessageMixin, DeleteView):
    model = Region
    success_url = reverse_lazy('region-list')
    success_message = 'Region "%(name)s" has been deleted successfully'

    def delete(self, request, *args, **kwargs):
        region = self.get_object()
        expense_type = SalesmanExpenseType.objects.filter(region=region).first()
        if expense_type:
            messages.error(
                self.request,
                'Cannot Delete region "{}" as it is still used by '
                'salesman "{}"'.format(region, expense_type.salesman)
            )
            return HttpResponseRedirect(reverse_lazy('region-list'))
        else:
            return super(RegionDelete, self).delete(request, *args, **kwargs)

